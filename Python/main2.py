import os
import tkinter as tk
import openpyxl

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("List new connection search")
        self.geometry("600x400")

        self.search_label = tk.Label(self, text="Enter S/C NO or MTR. NO:")
        self.search_label.pack(pady=10)

        self.search_entry = tk.Entry(self, font=("Arial", 14))
        self.search_entry.pack(pady=5)
        self.search_entry.bind("<Return>", lambda event: self.search_files())  # Bind search function to Enter key

        self.search_button = tk.Button(self, text="Search", command=self.search_files)
        self.search_button.pack(pady=10)

        self.result_label = tk.Label(self, text="Results:", font=("Arial", 14, "bold"))
        self.result_label.pack(pady=10)

        self.result_text = tk.Text(self, height=10, width=60, font=("Courier", 12))
        self.result_text.pack(pady=5)

    def search_files(self):
        search_id = self.search_entry.get().strip()  # Ensure search_id is stripped of leading/trailing spaces
        if not search_id:
            self.result_text.delete("1.0", tk.END)
            self.result_text.insert(tk.END, "Please enter an ID to search.")
            return

        self.result_text.delete("1.0", tk.END)
        drive = "D:\\"
        found_row = None  # Initialize found_row to None

        try:
            for root, dirs, files in os.walk(drive):
                for file in files:
                    # Only process files that end with.xlsx and start with 'List'
                    if file.lower().endswith('.xlsx') and file.lower().startswith('list'):
                        file_path = os.path.join(root, file)
                        workbook = openpyxl.load_workbook(file_path)
                        for sheet in workbook.sheetnames:
                            worksheet = workbook[sheet]
                            # Dynamically find columns that contain the search criteria
                            for col in worksheet.iter_cols():
                                column_values = [cell.value for cell in col]
                                if any(search_id in str(value) for value in column_values):
                                    # If found, iterate through rows to find the matching row
                                    for row in worksheet.iter_rows(values_only=True):
                                        if any(search_id in str(value) for value in row):
                                            # Filter out None values
                                            filtered_row = [value if value is not None else "" for value in row]
                                            found_row = (filtered_row, file_path)  # Store the found row and file path
                                            break  # Stop searching after finding a match
                                    if found_row:
                                        break  # Exit inner loop once a match is found
                            if found_row:
                                break  # Exit outer loop once a match is found
                    if found_row:
                        break  # Exit main loop once a match is found
        except PermissionError:
            self.result_text.insert(tk.END, "Permission denied. Please check your permissions.\n")
        except Exception as e:
            self.result_text.insert(tk.END, f"An error occurred: {str(e)}\n")

        if found_row:
            row, file_path = found_row
            # Define field names and values
            fields = ["SL NO", "Name", "Work Order No", "Work Order Date", "Application NO", "S/C NO", "Load", "Meter NO", "Wire Size", "Rt Lenght", "Meter Issued On", "Date of inespection", "Date of excecution", "Installation No"]
            values = row

            # Display field names and values
            for field, value in zip(fields, values):
                self.result_text.insert(tk.END, f"{field:<25}: {value}\n")
            
            # Append file name and path
            self.result_text.insert(tk.END, f"FILE NAME             : {os.path.basename(file_path)}\n")
            self.result_text.insert(tk.END, f"FILE PATH             : {file_path}\n")
            self.result_text.insert(tk.END, "\n")
        else:
            self.result_text.insert(tk.END, "No matching rows found.")

if __name__ == "__main__":
    app = App()
    app.mainloop()
