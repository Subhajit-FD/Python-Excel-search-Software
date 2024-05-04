import os
import tkinter as tk
from tkinter import filedialog, ttk
import openpyxl

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("New Service Connection File search")
        self.geometry("600x400")

        self.search_label = tk.Label(self, text="Enter S/C NO or MTR. NO:")
        self.search_label.pack(pady=10)

        self.search_entry = tk.Entry(self, font=("Arial", 14))
        self.search_entry.pack(pady=5)
        self.search_entry.bind("<Return>", lambda event: self.search_files())  # Bind search function to Enter key

        self.search_button = tk.Button(self, text="Search", command=self.search_files)
        self.search_button.pack(pady=10)

        self.result_label = tk.Label(self, text="Results:", font=("Arial", 14, "bold"))
        self.result_label.pack(pady=10)

        self.result_text = tk.Text(self, height=10, width=60, font=("Courier", 12))
        self.result_text.pack(pady=5)

    def search_files(self):
        search_id = self.search_entry.get().strip()  # Strip leading and trailing spaces from search_id
        if not search_id:
            self.result_text.delete("1.0", tk.END)
            self.result_text.insert(tk.END, "Please enter an ID to search.")
            return

        self.result_text.delete("1.0", tk.END)
        drive = "D:\\"
        found_rows = []

        try:
            for root, dirs, files in os.walk(drive):
                for file in files:
                    if file.lower().endswith(('.xlsx', '.xlsm')):
                        file_path = os.path.join(root, file)
                        workbook = openpyxl.load_workbook(file_path)
                        for sheet in workbook.sheetnames:
                            worksheet = workbook[sheet]
                            for row in worksheet.iter_rows(min_row=1, values_only=True):
                                # Strip spaces from each cell value before comparing
                                if search_id in [str(value).strip() for value in row[4:6]]:
                                    # Filter out None values
                                    filtered_row = [value if value is not None else "" for value in row]
                                    found_rows.append((filtered_row, file_path))
        except PermissionError:
            self.result_text.insert(tk.END, "Permission denied. Please check your permissions.\n")
        except Exception as e:
            self.result_text.insert(tk.END, f"An error occurred: {str(e)}\n")

        if found_rows:
            for row, file_path in found_rows:
                # Define field names and values
                fields = ["W.O NO", "DATE", "NAME", "APP NO", "S/C NO", "MTR. NO", "DT OF CON/ INSTALL NO", "INSTALL NO", "P.O NO & DATE", "SES NO", "REMARK"]
                values = row

                # Display field names and values
                for field, value in zip(fields, values):
                    self.result_text.insert(tk.END, f"{field:<25}: {value}\n")
                
                # Append file name and path
                self.result_text.insert(tk.END, f"FILE NAME             : {os.path.basename(file_path)}\n")
                self.result_text.insert(tk.END, f"FILE PATH             : {file_path}\n")
                self.result_text.insert(tk.END, "\n")
        else:
            self.result_text.insert(tk.END, "No matching rows found.")

if __name__ == "__main__":
    app = App()
    app.mainloop()
